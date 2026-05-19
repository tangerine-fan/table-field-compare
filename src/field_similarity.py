#!/usr/bin/env python3
"""
字段相似性分析工具
读取 table_field_compare.py 的输出，对未匹配字段（仅DEV有 / 仅标准化有）做相似度分析，
生成左右对照表，帮助判断是否为"同名不同译"的字段。

用法:
    python field_similarity.py 字段对比结果.xlsx -o 相似性分析.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

# ── 导入样式 ──
sys.path.insert(0, str(Path(__file__).resolve().parent))
from excel_style import (
    ALIGN_CENTER, ALIGN_LEFT, BODY_FONT, DETAIL_COL_WIDTHS,
    HEADER_FILL, HEADER_FONT, SOURCE_FILLS, SOURCE_FONTS, THIN_BORDER,
    freeze_header, set_column_widths, style_data_cell, style_header_row,
)

SIMILARITY_COL_WIDTHS = [10, 25, 25, 15, 25, 15, 10, 10]


# ── 相似度计算 ──

def cn_similarity(cn1: str, cn2: str) -> float:
    """中文名相似度 (0-1)"""
    if not cn1 or not cn2:
        return 0.0
    return SequenceMatcher(None, cn1, cn2).ratio()


def en_similarity(en1: str, en2: str) -> float:
    """英文名相似度 (0-1)：基于公共子序列"""
    if not en1 or not en2:
        return 0.0
    s1, s2 = en1.upper(), en2.upper()
    # 最长公共子串占比
    m = [[0] * (len(s2) + 1) for _ in range(len(s1) + 1)]
    longest = 0
    for i in range(len(s1)):
        for j in range(len(s2)):
            if s1[i] == s2[j]:
                m[i + 1][j + 1] = m[i][j] + 1
                longest = max(longest, m[i + 1][j + 1])
    if longest == 0:
        # 回退到字符级 SequenceMatcher
        return SequenceMatcher(None, s1, s2).ratio()
    return longest / max(len(s1), len(s2))


def combined_similarity(cn1: str, cn2: str, en1: str, en2: str) -> tuple[float, float, float]:
    """综合相似度，返回 (综合分, 中文分, 英文分)"""
    cs = cn_similarity(cn1, cn2)
    es = en_similarity(en1, en2)
    # 加权：中文 60% + 英文 40%
    combined = cs * 0.6 + es * 0.4
    return combined, cs, es


def confidence_level(score: float) -> str:
    """分数 → 置信度"""
    if score >= 0.85:
        return "高"
    elif score >= 0.6:
        return "中"
    else:
        return "低"


# ── 匹配逻辑 ──

def match_fields(dev_fields: list[dict], std_fields: list[dict]) -> list[dict]:
    """
    对 DEV 独有和标准化独有字段做交叉匹配。
    每个 DEV 字段找最佳匹配的标准化字段，并去重。
    """
    used_std: set[int] = set()
    pairs: list[dict] = []

    for di, dev in enumerate(dev_fields):
        best_score = 0.0
        best_cs = 0.0
        best_es = 0.0
        best_si = -1

        for si, std in enumerate(std_fields):
            if si in used_std:
                continue
            combined, cs, es = combined_similarity(
                dev["cn"], std["cn"], dev["en"], std["en"]
            )
            if combined > best_score:
                best_score = combined
                best_cs = cs
                best_es = es
                best_si = si

        if best_si >= 0:
            used_std.add(best_si)
            pairs.append({
                "dev_en": dev["en"], "dev_cn": dev["cn"],
                "std_en": std_fields[best_si]["en"],
                "std_cn": std_fields[best_si]["cn"],
                "cn_sim": round(best_cs * 100),
                "en_sim": round(best_es * 100),
                "combined": round(best_score * 100),
                "confidence": confidence_level(best_score),
            })
        else:
            pairs.append({
                "dev_en": dev["en"], "dev_cn": dev["cn"],
                "std_en": "", "std_cn": "",
                "cn_sim": 0, "en_sim": 0, "combined": 0,
                "confidence": "无匹配",
            })

    # 未匹配的标准化字段
    for si, std in enumerate(std_fields):
        if si not in used_std:
            pairs.append({
                "dev_en": "", "dev_cn": "",
                "std_en": std["en"], "std_cn": std["cn"],
                "cn_sim": 0, "en_sim": 0, "combined": 0,
                "confidence": "无匹配",
            })

    return pairs


# ── 读取 ──

def read_comparison(filepath: str) -> list[dict]:
    """读取字段对比结果 Sheet2，提取未匹配字段"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if "字段级详细对比" not in wb.sheetnames:
        raise ValueError(f"找不到'字段级详细对比' sheet，可用: {wb.sheetnames}")

    ws = wb["字段级详细对比"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header = rows[0]
    # 检测列位置（因为列可能被重排）
    col_map = {}
    for i, h in enumerate(header):
        if h == "模式":
            col_map["mode"] = i
        elif h == "标准表名":
            col_map["table"] = i
        elif h == "字段名":
            col_map["field"] = i
        elif h == "标准化中文名":
            col_map["std_cn"] = i
        elif h == "DEV中文名":
            col_map["dev_cn"] = i
        elif h == "字段来源":
            col_map["source"] = i

    # 按表分组收集未匹配字段
    tables: dict[tuple[str, str], dict[str, list[dict]]] = defaultdict(
        lambda: {"only_dev": [], "only_std": []}
    )

    for row in rows[1:]:
        if not row or not any(row):
            continue
        mode = str(row[col_map["mode"]]).strip() if row[col_map["mode"]] else ""
        table = str(row[col_map["table"]]).strip() if row[col_map["table"]] else ""
        field = str(row[col_map["field"]]).strip() if row[col_map["field"]] else ""
        source = str(row[col_map["source"]]).strip() if row[col_map["source"]] else ""
        std_cn = str(row[col_map["std_cn"]]).strip() if row[col_map["std_cn"]] else ""
        dev_cn = str(row[col_map["dev_cn"]]).strip() if row[col_map["dev_cn"]] else ""

        if not table or not field:
            continue

        key = (mode, table)
        if source == "仅DEV有":
            tables[key]["only_dev"].append({"en": field, "cn": dev_cn})
        elif source == "仅标准化有":
            tables[key]["only_std"].append({"en": field, "cn": std_cn})

    return tables


# ── 写入 ──

def write_similarity(tables: dict, output_path: str) -> None:
    """生成相似性分析 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "字段相似性分析"

    headers = ["模式", "标准表名", "DEV字段", "DEV中文名", "标准化字段", "标准化中文名", "相似度%", "置信度"]
    style_header_row(ws, headers)

    row_idx = 2
    for (mode, table), groups in sorted(tables.items()):
        dev_fields = groups["only_dev"]
        std_fields = groups["only_std"]

        if not dev_fields and not std_fields:
            continue

        pairs = match_fields(dev_fields, std_fields)

        for pair in pairs:
            ws.cell(row=row_idx, column=1, value=mode)
            ws.cell(row=row_idx, column=2, value=table)
            ws.cell(row=row_idx, column=3, value=pair["dev_en"])
            ws.cell(row=row_idx, column=4, value=pair["dev_cn"])
            ws.cell(row=row_idx, column=5, value=pair["std_en"])
            ws.cell(row=row_idx, column=6, value=pair["std_cn"])
            ws.cell(row=row_idx, column=7, value=pair["combined"])
            ws.cell(row=row_idx, column=8, value=pair["confidence"])

            # 样式
            conf = pair["confidence"]
            font = SOURCE_FONTS.get("两边共有", BODY_FONT) if conf == "高" else BODY_FONT
            for col in range(1, 9):
                fill = None
                if conf == "高" and col == 8:
                    fill = SOURCE_FILLS["两边共有"]
                elif conf == "低" and col == 8:
                    fill = SOURCE_FILLS["仅标准化有"]
                style_data_cell(ws, row_idx, col, font=font, fill=fill)

            row_idx += 1

    set_column_widths(ws, SIMILARITY_COL_WIDTHS)
    freeze_header(ws)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"✅ 相似性分析已保存: {output_path}")
    print(f"   共 {row_idx - 2} 行")


# ── 入口 ──

def main():
    parser = argparse.ArgumentParser(description="字段相似性分析工具")
    parser.add_argument("input", help="table_field_compare.py 的输出 Excel")
    parser.add_argument("-o", "--output", default="字段相似性分析.xlsx", help="输出路径")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"错误: 文件不存在: {args.input}")
        sys.exit(1)

    print(f"读取: {args.input}")
    tables = read_comparison(args.input)

    total_tables = len(tables)
    total_dev = sum(len(v["only_dev"]) for v in tables.values())
    total_std = sum(len(v["only_std"]) for v in tables.values())
    print(f"  {total_tables} 个表, {total_dev} DEV独有字段, {total_std} 标准化独有字段")

    write_similarity(tables, args.output)


if __name__ == "__main__":
    main()
