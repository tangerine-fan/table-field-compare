"""Generated Excel style: hermes_blue
企业蓝表头 + 绿/红/黄状态色 — 稳重现代（默认）

此文件由 Tools/excel_styles/generate.py 生成，独立可用，只依赖 openpyxl。
如需更换风格，重新运行生成器即可。
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 主题信息 ──
NAME = "hermes_blue"
DESCRIPTION = "企业蓝表头 + 绿/红/黄状态色 — 稳重现代（默认）"

# ── 字体 ──
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="00FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=11)
SUMMARY_FONT = Font(name="微软雅黑", bold=True, size=11)

# ── 填充 ──
HEADER_FILL = PatternFill(start_color="004472C4", end_color="004472C4", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="00D9E2F3", end_color="00D9E2F3", fill_type="solid")

STATUS_FILLS = {
    "完全一致": PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid"),
    "存在差异": PatternFill(start_color="00FFC7CE", end_color="00FFC7CE", fill_type="solid"),
    "DEV缺失": PatternFill(start_color="00FFEB9C", end_color="00FFEB9C", fill_type="solid"),
    "标准化缺失": PatternFill(start_color="00FFEB9C", end_color="00FFEB9C", fill_type="solid"),
}

STATUS_FONTS = {
    "完全一致": Font(name="微软雅黑", color="00006100"),
    "存在差异": Font(name="微软雅黑", color="009C0006"),
    "DEV缺失": Font(name="微软雅黑", color="009C6500"),
    "标准化缺失": Font(name="微软雅黑", color="009C6500"),
}

SOURCE_FILLS = {
    "两边共有": PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid"),
    "仅标准化有": PatternFill(start_color="00FFC7CE", end_color="00FFC7CE", fill_type="solid"),
    "仅DEV有": PatternFill(start_color="00FFEB9C", end_color="00FFEB9C", fill_type="solid"),
}

SOURCE_FONTS = {
    "两边共有": Font(name="微软雅黑", color="00006100"),
    "仅标准化有": Font(name="微软雅黑", color="009C0006"),
    "仅DEV有": Font(name="微软雅黑", color="009C6500"),
}

# ── 对齐 ──
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)

# ── 边框 ──
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── 列宽 ──
SUMMARY_COL_WIDTHS = [25, 35, 15, 15, 15, 80]
DETAIL_COL_WIDTHS = [25, 35, 30, 15]


# ── 辅助函数 ──

def style_header_row(ws, headers):
    """为工作表设置标准表头行（第1行）"""
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER


def style_data_cell(ws, row, col, font=None, fill=None):
    """为单个数据单元格设置字体、边框、对齐"""
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    cell.font = font or BODY_FONT
    cell.alignment = ALIGN_LEFT
    if fill:
        cell.fill = fill


def set_column_widths(ws, widths):
    """批量设置列宽（1-indexed）"""
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_header(ws):
    """冻结首行（表头）"""
    ws.freeze_panes = "A2"


def style_summary_row(ws, row, col_count, text):
    """设置统计汇总行"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUMMARY_FONT
    cell.fill = SUMMARY_FILL
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    for col_idx in range(2, col_count + 1):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER


# ── 跨平台字体检测（导入时自动执行） ──

def _detect_cjk_font():
    """检测系统可用的中文字体，按优先级回退。"""
    import platform
    import subprocess

    system = platform.system()
    if system == "Windows":
        return "微软雅黑"
    elif system == "Darwin":
        return "PingFang SC"
    else:
        try:
            result = subprocess.run(
                ["fc-list", ":lang=zh", "-f", "%{family}\n"],
                capture_output=True, text=True, timeout=3,
            )
            installed = set(result.stdout.strip().split("\n"))
            for candidate in [
                "微软雅黑", "Microsoft YaHei", "SimHei", "SimSun",
                "Noto Sans CJK SC", "Noto Sans CJK TC",
                "WenQuanYi Micro Hei", "Droid Sans Fallback",
            ]:
                if candidate in installed:
                    return candidate
        except Exception:
            pass
        return "Noto Sans CJK SC"


_detected_font = _detect_cjk_font()
for _f in [HEADER_FONT, BODY_FONT, SUMMARY_FONT,
           *STATUS_FONTS.values(), *SOURCE_FONTS.values()]:
    _f.name = _detected_font
