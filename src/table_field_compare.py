#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
表结构字段对比工具
对比TEST_DEV与标准化文件的字段数量和字段名差异

使用方式:
    python table_field_compare.py <TEST_DEV.xlsx> <standard.xlsx> [-o output.xlsx]

    # 仅控制台输出
    python table_field_compare.py dev.xlsx std.xlsx --console-only

    # 自定义排除字段
    python table_field_compare.py dev.xlsx std.xlsx --exclude DEL_FLG,PART_DT

    # 指定输出路径
    python table_field_compare.py dev.xlsx std.xlsx -o result.xlsx

作者: AI Assistant
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl

# ── 导入 Excel 样式工具（workspace 层 Tools/excel_styles） ──
_workspace_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
if _workspace_root not in sys.path:
    sys.path.insert(0, _workspace_root)

from Tools.excel_styles import (
    freeze_header,
    get_theme,
    set_column_widths,
    style_data_cell,
    style_header_row,
    style_summary_row,
    update_font_name,
)

# ========== 日志配置 ==========

logging.basicConfig(
    level=logging.INFO,
    format="[%(levelname)s] %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger(__name__)


# ========== 配置 ==========

# 默认需要从DEV中排除的字段（不参与对比，视为不存在）
DEFAULT_EXCLUDE_FIELDS = {"DEL_FLG", "ETL_TM_STMP", "PART_DT", "SRC_SYS_CD"}

# 数据行搜索上限（防止死循环）
MAX_DATA_SEARCH_ROWS = 200

# ========== 字体自动检测 ==========


def _detect_cjk_font() -> str:
    """
    检测系统可用的中文字体，按优先级回退。
    优先级: 微软雅黑 > SimHei > SimSun > Noto Sans CJK SC > Droid Sans Fallback > 系统默认
    """
    _FONT_CANDIDATES = [
        "微软雅黑",
        "Microsoft YaHei",
        "SimHei",
        "SimSun",
        "Noto Sans CJK SC",
        "Noto Sans CJK TC",
        "WenQuanYi Micro Hei",
        "WenQuanYi Zen Hei",
        "Droid Sans Fallback",
        "Arial Unicode MS",
    ]
    # openpyxl 不提供字体检测 API，这里按平台策略选择
    import platform

    system = platform.system()
    if system == "Windows":
        # Windows 通常有微软雅黑
        return "微软雅黑"
    elif system == "Darwin":
        return "PingFang SC"
    else:
        # Linux: 尝试 fc-list 检测
        import subprocess

        try:
            result = subprocess.run(
                ["fc-list", ":lang=zh", "-f", "%{family}\n"],
                capture_output=True, text=True, timeout=3,
            )
            installed = set(result.stdout.strip().split("\n"))
            for candidate in _FONT_CANDIDATES:
                if candidate in installed:
                    logger.debug("检测到中文字体: %s", candidate)
                    return candidate
        except Exception:
            pass
        # 回退
        return "Noto Sans CJK SC"


FONT_NAME = _detect_cjk_font()
logger.info("使用字体: %s", FONT_NAME)

# ========== 加载主题 ==========

# 加载默认主题 (hermes_blue)，应用系统检测到的字体
_theme = get_theme()
update_font_name(_theme, FONT_NAME)
logger.info("使用样式主题: %s", _theme.NAME)


# ========== 核心函数 ==========


def read_test_dev(filepath: str) -> dict[str, list[str]]:
    """
    读取TEST_DEV文件，从"字段信息"sheet中提取表名和字段名。
    表头包含"SCHEMA标识"和"表名"，数据从表头下一行开始。
    返回: {原始DEV表名: [字段列表]}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if "字段信息" not in wb.sheetnames:
        raise ValueError(
            f"TEST_DEV文件中找不到'字段信息' sheet。"
            f"可用 sheet: {', '.join(wb.sheetnames)}"
        )

    ws = wb["字段信息"]
    rows = list(ws.iter_rows(values_only=True))

    # 找到包含"SCHEMA标识"或"表名"的表头行
    header_row_idx: Optional[int] = None
    for i, row in enumerate(rows):
        if row[0] == "SCHEMA标识" or (len(row) > 1 and row[1] == "表名"):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("无法找到TEST_DEV的表头行（应包含'SCHEMA标识'和'表名'）")

    header = rows[header_row_idx]

    # 定位表名列和字段名列
    table_idx: Optional[int] = None
    field_idx: Optional[int] = None
    for i, h in enumerate(header):
        if h == "表名":
            table_idx = i
        elif h == "字段名":
            field_idx = i

    if table_idx is None or field_idx is None:
        raise ValueError(f"无法定位'表名'或'字段名'列，表头: {header[:5]}")

    # 从表头下一行开始，跳过说明行，最多扫描 MAX_DATA_SEARCH_ROWS 行
    data_start = header_row_idx + 1
    max_search = min(data_start + MAX_DATA_SEARCH_ROWS, len(rows))
    while data_start < max_search:
        row = rows[data_start]
        if row[table_idx] is not None or row[field_idx] is not None:
            break
        data_start += 1
    else:
        logger.warning(
            "在表头后 %d 行内未找到数据行，尝试从表头+1行开始读取",
            MAX_DATA_SEARCH_ROWS,
        )
        data_start = header_row_idx + 1

    # 收集数据
    tables: dict[str, list[str]] = defaultdict(list)
    for row in rows[data_start:]:
        table = str(row[table_idx]).strip() if row[table_idx] else None
        field = str(row[field_idx]).strip() if row[field_idx] else None
        if table and field:
            tables[table].append(field)

    return dict(tables)


def is_field_name(value) -> bool:
    """
    判断一个值是否像字段名。
    支持大写下划线格式（如 OVLMT_DATE）和中文命名字段。
    """
    if not value or not isinstance(value, str):
        return False
    v = value.strip()
    if not v:
        return False
    # 大写下划线格式: OVLMT_DATE, ACCT_NO
    if "_" in v and v.replace("_", "").isalnum() and v.isupper():
        return True
    # 全大写字母: OVLMTDATE
    if v.isalpha() and v.isupper() and 2 <= len(v) <= 30:
        return True
    # 中文命名字段（包含中文字符且不以数字开头）
    if any("\u4e00" <= ch <= "\u9fff" for ch in v) and not v[0].isdigit():
        return True
    return False


def read_standard_file(filepath: str) -> dict[str, list[str]]:
    """
    读取标准化文件，智能检测字段名所在的列。
    表头包含"表名"，数据从表头下一行开始。
    返回: {标准表名: [字段列表]}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # 找到包含"表名"的表头行
    header_row_idx: Optional[int] = None
    for i, row in enumerate(rows):
        for cell in row:
            if cell and "表名" in str(cell):
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError("无法找到标准化文件的表头行（应包含'表名'）")

    header = rows[header_row_idx]

    # 找表名列
    table_idx: Optional[int] = None
    for i, h in enumerate(header):
        if h and "表名" in str(h):
            table_idx = i
            break

    if table_idx is None:
        raise ValueError("无法定位'表名'列")

    # 智能检测字段列
    field_idx: Optional[int] = None

    # 采样行：表头后最多5行
    sample_end = min(header_row_idx + 6, len(rows))
    sample_rows = rows[header_row_idx + 1 : sample_end]

    # 策略1: 从列头文字找"字段名称"
    for i, h in enumerate(header):
        if h and "字段名称" in str(h):
            for sr in sample_rows:
                if sr[i] and is_field_name(sr[i]):
                    field_idx = i
                    break
            if field_idx is not None:
                break

    # 策略2: 用数据特征检测前5个列
    if field_idx is None:
        best_col: Optional[int] = None
        best_score = 0
        for col_idx in range(min(5, len(header))):
            if col_idx == table_idx:
                continue
            score = 0
            for sr in sample_rows:
                if sr[col_idx] and is_field_name(sr[col_idx]):
                    score += 2
            if score > best_score:
                best_score = score
                best_col = col_idx

        if best_col is not None and best_score > 0:
            field_idx = best_col
        else:
            # 如果没有任何数据行，返回空表（文件只有表头）
            if len(sample_rows) == 0 or all(all(c is None for c in sr) for sr in sample_rows):
                logger.warning("标准化文件无数据行，返回空表")
                return {}
            raise ValueError(
                f"无法定位字段列，请检查文件格式。\n"
                f"表头: {list(header[:8])}\n"
                f"采样数据: {[list(sr[:5]) if sr else [] for sr in sample_rows[:2]]}"
            )

    logger.info("标准化文件: 表名在列%d，字段名在列%d", table_idx, field_idx)

    # 收集数据
    tables: dict[str, list[str]] = defaultdict(list)
    for row in rows[header_row_idx + 1 :]:
        table = str(row[table_idx]).strip() if row[table_idx] else None
        field = str(row[field_idx]).strip() if row[field_idx] else None
        if table and field:
            tables[table].append(field)

    return dict(tables)


def extract_std_name(dev_table_name: str) -> str:
    """
    从TEST_DEV表名提取标准表名。
    格式: S_系统名_标准中表名_接入方式
    例:  S_CCS_CCS_CARD_ACCT_MALL  ->  CCS_CARD_ACCT
    """
    if not dev_table_name or not isinstance(dev_table_name, str):
        return dev_table_name

    parts = dev_table_name.split("_")
    if len(parts) >= 4 and parts[0] == "S":
        extracted = "_".join(parts[2:-1])  # 去掉 S_系统名_ 和 _接入方式
        if not extracted:
            logger.warning(
                "表名 '%s' 按规则提取后为空，保留原名", dev_table_name
            )
            return dev_table_name
        return extracted

    # 不符合标准格式，保留原名
    logger.debug(
        "表名 '%s' 不符合 S_系统名_标准名_接入方式 格式，保留原名", dev_table_name
    )
    return dev_table_name


def compare_fields(
    dev_tables_raw: dict[str, list[str]],
    std_tables: dict[str, list[str]],
    exclude_fields: set[str],
) -> list[dict]:
    """
    对比两边的字段，返回对比结果列表。
    自动排除 exclude_fields 中指定的字段。
    """
    # 提取DEV标准表名，并过滤排除字段
    dev_mapping: dict[str, dict] = {}
    for dev_name, fields in dev_tables_raw.items():
        std_name = extract_std_name(dev_name)
        filtered_fields = [f for f in fields if f not in exclude_fields]
        dev_mapping[std_name] = {"dev_name": dev_name, "fields": filtered_fields}

    # 取两边表名的并集
    all_tables = sorted(set(dev_mapping.keys()) | set(std_tables.keys()))

    results: list[dict] = []
    for std_name in all_tables:
        dev_info = dev_mapping.get(std_name)
        std_fields_list = std_tables.get(std_name)

        dev_name = dev_info["dev_name"] if dev_info else None
        dev_field_count = len(dev_info["fields"]) if dev_info else 0
        dev_field_set = set(dev_info["fields"]) if dev_info else set()

        std_field_count = len(std_fields_list) if std_fields_list else 0
        std_field_set = set(std_fields_list) if std_fields_list else set()

        has_dev = dev_info is not None
        has_std = std_fields_list is not None

        common = std_field_set & dev_field_set
        only_std = std_field_set - dev_field_set
        only_dev = dev_field_set - std_field_set

        # 判断对比结果
        if not has_dev:
            status = "DEV缺失"
            diff_desc = f"标准化有{std_field_count}个字段，DEV中无此表"
        elif not has_std:
            status = "标准化缺失"
            diff_desc = f"DEV有{dev_field_count}个字段，标准化中无此表"
        elif len(only_std) == 0 and len(only_dev) == 0:
            status = "完全一致"
            diff_desc = f"字段数量和字段名完全一致（{dev_field_count}个字段）"
        else:
            status = "存在差异"
            parts = []
            if dev_field_count != std_field_count:
                parts.append(
                    f"字段数不同（标准化{std_field_count} vs DEV{dev_field_count}）"
                )
            if only_std:
                parts.append(
                    f"标准化独有{len(only_std)}个：{', '.join(sorted(only_std))}"
                )
            if only_dev:
                parts.append(
                    f"DEV独有{len(only_dev)}个：{', '.join(sorted(only_dev))}"
                )
            diff_desc = " | ".join(parts)

        results.append(
            {
                "标准表名": std_name,
                "DEV原表名": dev_name or "(缺失)",
                "标准化字段数": std_field_count,
                "DEV字段数": dev_field_count,
                "对比结果": status,
                "差异说明": diff_desc,
                "_common": sorted(common),
                "_only_std": sorted(only_std),
                "_only_dev": sorted(only_dev),
            }
        )

    return results


def write_excel(results: list[dict], output_path: str) -> None:
    """将对比结果写入Excel文件，包含两个Sheet"""
    theme = get_theme()  # hermes_blue

    wb = openpyxl.Workbook()

    # ----- Sheet 1: 汇总对比 -----
    ws1 = wb.active
    ws1.title = "汇总对比"
    headers1 = ["标准表名", "DEV原表名", "标准化字段数", "DEV字段数", "对比结果", "差异说明"]

    style_header_row(ws1, headers1, theme)

    # 批量收集数据行
    data_rows: list[list] = []
    for r in results:
        data_rows.append([
            r["标准表名"], r["DEV原表名"], r["标准化字段数"],
            r["DEV字段数"], r["对比结果"], r["差异说明"],
        ])

    # 写数据行 + 样式
    for r_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=r_idx, column=col_idx, value=value)

        status = row_data[4]
        status_font = theme.STATUS_FONTS.get(status, theme.BODY_FONT)
        status_fill = theme.STATUS_FILLS.get(status)

        for col_idx in range(1, 7):
            style_data_cell(ws1, r_idx, col_idx, font=status_font, fill=status_fill if col_idx == 5 else None, theme=theme)

    # 统计汇总行
    summary_row = len(data_rows) + 2
    total = len(results)
    perfect = sum(1 for r in results if r["对比结果"] == "完全一致")
    diff = sum(1 for r in results if r["对比结果"] == "存在差异")
    dev_miss = sum(1 for r in results if r["对比结果"] == "DEV缺失")
    std_miss = sum(1 for r in results if r["对比结果"] == "标准化缺失")
    summary_text = (
        f"共 {total} 个表 | "
        f"完全一致: {perfect} | "
        f"存在差异: {diff} | "
        f"DEV缺失: {dev_miss} | "
        f"标准化缺失: {std_miss}"
    )
    style_summary_row(ws1, summary_row, 6, summary_text, theme)

    set_column_widths(ws1, theme.SUMMARY_COL_WIDTHS)
    freeze_header(ws1)

    # ----- Sheet 2: 字段级详细对比 -----
    ws2 = wb.create_sheet("字段级详细对比")
    headers2 = ["标准表名", "DEV原表名", "字段名", "字段来源"]
    style_header_row(ws2, headers2, theme)

    # 批量收集所有字段行
    field_rows: list[tuple[str, str, str, str]] = []
    for r in results:
        for field in r["_common"]:
            field_rows.append((r["标准表名"], r["DEV原表名"], field, "两边共有"))
        for field in r["_only_std"]:
            field_rows.append((r["标准表名"], r["DEV原表名"], field, "仅标准化有"))
        for field in r["_only_dev"]:
            field_rows.append((r["标准表名"], r["DEV原表名"], field, "仅DEV有"))

    for r_idx, (table, dev_name, field, source) in enumerate(field_rows, 2):
        ws2.cell(row=r_idx, column=1, value=table)
        ws2.cell(row=r_idx, column=2, value=dev_name)
        ws2.cell(row=r_idx, column=3, value=field)
        ws2.cell(row=r_idx, column=4, value=source)

        source_font = theme.SOURCE_FONTS.get(source, theme.BODY_FONT)
        source_fill = theme.SOURCE_FILLS.get(source)
        for col_idx in range(1, 5):
            style_data_cell(ws2, r_idx, col_idx, font=source_font,
                           fill=source_fill if col_idx == 4 else None, theme=theme)

    set_column_widths(ws2, theme.DETAIL_COL_WIDTHS)
    freeze_header(ws2)

    # 保存
    output_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    logger.info("✅ 对比结果已保存: %s", output_path)


def print_summary(results: list[dict]) -> None:
    """在控制台打印对比结果摘要"""
    print("\n" + "=" * 70)
    print("对比结果摘要")
    print("=" * 70)

    for r in results:
        print(f"\n  表: {r['标准表名']} ({r['DEV原表名']})")
        print(f"    标准化字段数: {r['标准化字段数']}")
        print(f"    DEV字段数（已排除指定字段）: {r['DEV字段数']}")
        print(f"    对比结果: {r['对比结果']}")

        if r["对比结果"] == "存在差异":
            print(f"    两边共有: {len(r['_common'])} 个")
            print(f"    标准化独有: {len(r['_only_std'])} 个 -> {r['_only_std']}")
            print(f"    DEV独有: {len(r['_only_dev'])} 个 -> {r['_only_dev']}")

        print(f"    差异说明: {r['差异说明']}")

    # 统计汇总
    total = len(results)
    perfect = sum(1 for r in results if r["对比结果"] == "完全一致")
    diff = sum(1 for r in results if r["对比结果"] == "存在差异")
    dev_miss = sum(1 for r in results if r["对比结果"] == "DEV缺失")
    std_miss = sum(1 for r in results if r["对比结果"] == "标准化缺失")

    print("\n" + "=" * 70)
    print(
        f"📊 统计: 共 {total} 表 | "
        f"完全一致: {perfect} | 存在差异: {diff} | "
        f"DEV缺失: {dev_miss} | 标准化缺失: {std_miss}"
    )
    print("=" * 70)


def parse_args() -> argparse.Namespace:
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description="表结构字段对比工具 — 对比 TEST_DEV 与标准化文件的字段差异",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s TEST_DEV.xlsx standard.xlsx
  %(prog)s TEST_DEV.xlsx standard.xlsx -o result.xlsx
  %(prog)s TEST_DEV.xlsx standard.xlsx --console-only
  %(prog)s TEST_DEV.xlsx standard.xlsx --exclude DEL_FLG,PART_DT,SRC_SYS_CD
        """,
    )

    parser.add_argument(
        "dev_path",
        metavar="TEST_DEV.xlsx",
        nargs="?",
        help="TEST_DEV 文件路径",
    )
    parser.add_argument(
        "std_path",
        metavar="standard.xlsx",
        nargs="?",
        help="标准化映射文件路径",
    )
    parser.add_argument(
        "-o",
        "--output",
        dest="output_path",
        default="表结构字段对比结果.xlsx",
        help="输出文件路径（默认: 表结构字段对比结果.xlsx）",
    )
    parser.add_argument(
        "--exclude",
        default=None,
        help="需要排除的字段，逗号分隔（默认: DEL_FLG,ETL_TM_STMP,PART_DT,SRC_SYS_CD）",
    )
    parser.add_argument(
        "--console-only",
        action="store_true",
        help="仅打印控制台输出，不生成 Excel 文件",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="显示详细调试信息",
    )

    return parser.parse_args()


# ========== 主入口 ==========


def main() -> None:
    args = parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)

    dev_path: str = args.dev_path
    std_path: str = args.std_path
    output_path: str = args.output_path
    console_only: bool = args.console_only

    # 解析排除字段
    if args.exclude is not None:
        exclude_fields = {f.strip() for f in args.exclude.split(",") if f.strip()}
    else:
        exclude_fields = DEFAULT_EXCLUDE_FIELDS

    # 检查文件存在
    if not Path(dev_path).exists():
        logger.error("文件不存在: %s", dev_path)
        sys.exit(1)
    if not Path(std_path).exists():
        logger.error("文件不存在: %s", std_path)
        sys.exit(1)

    logger.info("TEST_DEV文件: %s", dev_path)
    logger.info("标准化文件: %s", std_path)
    logger.info("输出文件: %s", output_path)
    logger.info("排除字段: %s", ", ".join(sorted(exclude_fields)))

    # 读取数据
    logger.info("读取TEST_DEV...")
    dev_tables = read_test_dev(dev_path)
    logger.info("  -> %d 个表", len(dev_tables))

    logger.info("读取标准化文件...")
    std_tables = read_standard_file(std_path)
    logger.info("  -> %d 个表", len(std_tables))

    # 对比
    logger.info("执行字段对比...")
    results = compare_fields(dev_tables, std_tables, exclude_fields)

    # 输出
    if not console_only:
        write_excel(results, output_path)
    print_summary(results)


if __name__ == "__main__":
    main()
