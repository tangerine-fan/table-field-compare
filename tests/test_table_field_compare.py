#!/usr/bin/env python3
"""
table_field_compare 测试套件
用法:
    python tests/test_table_field_compare.py
    python tests/test_table_field_compare.py -v   # 详细
"""
from __future__ import annotations

import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

# ── 定位脚本 ──
PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRIPT = PROJECT_ROOT / "src" / "table_field_compare.py"

# ── 测试数据工厂 ──


def make_test_dev(path: str, tables: dict[str, list[str]]) -> str:
    """创建 TEST_DEV 格式的 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "字段信息"
    ws.cell(row=1, column=1, value="SCHEMA标识")
    ws.cell(row=1, column=2, value="表名")
    ws.cell(row=1, column=3, value="字段名")
    row_idx = 2
    for table, fields in tables.items():
        for field in fields:
            ws.cell(row=row_idx, column=2, value=table)
            ws.cell(row=row_idx, column=3, value=field)
            row_idx += 1
    wb.save(path)
    return path


def make_std_file(path: str, tables: dict[str, list[str]]) -> str:
    """创建标准化映射 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "标准化字段映射"
    ws.cell(row=1, column=1, value="标准化字段映射表")
    ws.cell(row=2, column=2, value="表名")
    ws.cell(row=2, column=3, value="字段名称")
    row_idx = 3
    for table, fields in tables.items():
        for field in fields:
            ws.cell(row=row_idx, column=1, value=row_idx - 2)
            ws.cell(row=row_idx, column=2, value=table)
            ws.cell(row=row_idx, column=3, value=field)
            row_idx += 1
    wb.save(path)
    return path


def run(*args: str, expect_ok: bool = True) -> subprocess.CompletedProcess:
    """运行脚本并检查退出码"""
    result = subprocess.run(
        [sys.executable, str(SCRIPT)] + list(args),
        capture_output=True, text=True, timeout=30,
    )
    if expect_ok and result.returncode != 0:
        raise AssertionError(
            f"脚本返回 {result.returncode}\nSTDERR: {result.stderr[:500]}"
        )
    return result


# ── 测试用例 ──


class TestCoreFunctionality(unittest.TestCase):
    """核心功能测试"""

    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="tfc_test_")

        # 正常数据
        cls.dev_path = os.path.join(cls.tmpdir, "dev.xlsx")
        make_test_dev(cls.dev_path, {
            "S_CCS_CCS_CARD_ACCT_MALL": ["ACCT_NO", "CARD_NO", "OVLMT_DATE", "OPN_DT", "DEL_FLG"],
            "S_CRD_CRD_CUST_INFO_MTH": ["CUST_ID", "CUST_NAME", "PART_DT"],
        })

        cls.std_path = os.path.join(cls.tmpdir, "std.xlsx")
        make_std_file(cls.std_path, {
            "CCS_CARD_ACCT": ["ACCT_NO", "CARD_NO", "OVLMT_DATE", "CARD_STS"],
            "CRD_CUST_INFO": ["CUST_ID", "CUST_NAME", "BIRTH_DT"],
        })

        # 空 DEV
        cls.empty_dev = os.path.join(cls.tmpdir, "empty_dev.xlsx")
        make_test_dev(cls.empty_dev, {})

        cls.empty_std = os.path.join(cls.tmpdir, "empty_std.xlsx")
        make_std_file(cls.empty_std, {})

    def test_normal_run(self):
        """正常对比生成 Excel"""
        output = os.path.join(self.tmpdir, "result.xlsx")
        result = run(str(self.dev_path), str(self.std_path), "-o", output)
        self.assertIn("统计", result.stdout)
        self.assertTrue(os.path.exists(output))

    def test_output_content(self):
        """验证输出 Excel 内容正确"""
        output = os.path.join(self.tmpdir, "result_content.xlsx")
        run(str(self.dev_path), str(self.std_path), "-o", output)

        wb = openpyxl.load_workbook(output)
        self.assertEqual(wb.sheetnames, ["汇总对比", "字段级详细对比"])

        ws1 = wb["汇总对比"]
        # 表头
        self.assertEqual(ws1.cell(row=1, column=1).value, "标准表名")
        # CCS_CARD_ACCT: 存在差异
        self.assertEqual(ws1.cell(row=2, column=1).value, "CCS_CARD_ACCT")
        self.assertEqual(ws1.cell(row=2, column=5).value, "存在差异")
        # CRD_CUST_INFO: 存在差异
        self.assertEqual(ws1.cell(row=3, column=1).value, "CRD_CUST_INFO")

    def test_exclude_fields(self):
        """排除字段过滤"""
        output = os.path.join(self.tmpdir, "result_exclude.xlsx")
        run(str(self.dev_path), str(self.std_path), "-o", output,
            "--exclude", "DEL_FLG,PART_DT")
        wb = openpyxl.load_workbook(output)
        ws1 = wb["汇总对比"]
        # CCS_CARD_ACCT DEV字段数应该是 4（排除了 DEL_FLG）
        self.assertEqual(ws1.cell(row=2, column=4).value, 4)

    def test_console_only(self):
        """--console-only 不生成文件"""
        output = os.path.join(self.tmpdir, "should_not_exist.xlsx")
        run(str(self.dev_path), str(self.std_path), "-o", output, "--console-only")
        self.assertFalse(os.path.exists(output))

    def test_empty_dev(self):
        """空 DEV 文件"""
        output = os.path.join(self.tmpdir, "result_empty_dev.xlsx")
        result = run(str(self.empty_dev), str(self.std_path), "-o", output)
        self.assertIn("DEV缺失", result.stdout)

    def test_empty_std(self):
        """空标准文件，DEV 有数据"""
        output = os.path.join(self.tmpdir, "result_empty_std.xlsx")
        result = run(str(self.dev_path), str(self.empty_std), "-o", output)
        self.assertIn("标准化缺失", result.stdout)

    def test_file_not_found(self):
        """文件不存在应报错"""
        result = run(
            "/tmp/nonexistent_12345.xlsx",
            str(self.std_path),
            "-o", os.path.join(self.tmpdir, "nope.xlsx"),
            expect_ok=False,
        )
        self.assertNotEqual(result.returncode, 0)

    def test_style_hermes_blue(self):
        """输出使用 hermes_blue 主题"""
        output = os.path.join(self.tmpdir, "style_check.xlsx")
        run(str(self.dev_path), str(self.std_path), "-o", output)
        wb = openpyxl.load_workbook(output)
        fill = wb["汇总对比"].cell(row=1, column=1).fill.start_color.rgb
        self.assertEqual(fill, "004472C4")


class TestCLI(unittest.TestCase):
    """命令行界面测试"""

    def test_help(self):
        result = run("--help")
        self.assertIn("--exclude", result.stdout)
        self.assertIn("--console-only", result.stdout)
        self.assertNotIn("--theme", result.stdout)  # 主题不可运行时切换


if __name__ == "__main__":
    unittest.main(verbosity=2)
