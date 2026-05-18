"""
openpyxl 写入辅助函数
提供表头、数据行、列宽、冻结窗格等常用 Excel 格式化操作。
"""
from __future__ import annotations

import importlib
from typing import TYPE_CHECKING, Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    from openpyxl.styles import Font, PatternFill

# 默认主题
_DEFAULT_THEME = "hermes_blue"

# 主题缓存
_theme_cache: dict[str, Any] = {}


def _load_theme(name: str):
    """惰性加载主题模块"""
    if name not in _theme_cache:
        # 优先相对导入（项目内），回退绝对导入（Tools/excel_styles）
        try:
            _theme_cache[name] = importlib.import_module(f".themes.{name}", package="excel_styles")
        except ImportError:
            _theme_cache[name] = importlib.import_module(f"Tools.excel_styles.themes.{name}")
    return _theme_cache[name]


def list_themes() -> list[dict[str, str]]:
    """列出所有可用主题"""
    names = [
        "hermes_blue",
        "obsidian_dark",
        "stripe_clean",
        "catppuccin_latte",
        "anuppuccin_warm",
    ]
    result = []
    for name in names:
        try:
            mod = _load_theme(name)
            result.append({"name": mod.NAME, "description": mod.DESCRIPTION})
        except ImportError:
            pass
    return result


def get_theme(name: str | None = None):
    """
    获取主题模块。不传则返回默认主题 (hermes_blue)。

    >>> theme = get_theme()
    >>> theme = get_theme("obsidian_dark")
    """
    return _load_theme(name or _DEFAULT_THEME)


def style_header_row(ws: Worksheet, headers: list[str], theme=None) -> None:
    """为工作表设置标准表头行（第1行）"""
    if theme is None:
        theme = get_theme()
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = theme.HEADER_FONT
        cell.fill = theme.HEADER_FILL
        cell.alignment = theme.ALIGN_CENTER


def style_data_cell(
    ws: Worksheet,
    row: int,
    col: int,
    font: Font | None = None,
    fill: PatternFill | None = None,
    theme=None,
) -> None:
    """为单个数据单元格设置字体、边框、对齐"""
    if theme is None:
        theme = get_theme()
    cell = ws.cell(row=row, column=col)
    cell.border = theme.THIN_BORDER
    cell.font = font or theme.BODY_FONT
    cell.alignment = theme.ALIGN_LEFT
    if fill:
        cell.fill = fill


def set_column_widths(ws: Worksheet, widths: list[int]) -> None:
    """批量设置列宽（1-indexed）"""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_header(ws: Worksheet) -> None:
    """冻结首行（表头）"""
    ws.freeze_panes = "A2"


def style_summary_row(
    ws: Worksheet,
    row: int,
    col_count: int,
    text: str,
    theme=None,
) -> None:
    """设置统计汇总行"""
    if theme is None:
        theme = get_theme()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = theme.SUMMARY_FONT
    cell.fill = theme.SUMMARY_FILL
    cell.alignment = theme.ALIGN_LEFT
    cell.border = theme.THIN_BORDER
    for col_idx in range(2, col_count + 1):
        ws.cell(row=row, column=col_idx).border = theme.THIN_BORDER


def update_font_name(theme, font_name: str):
    """
    运行时替换主题字体名（用于跨平台字体适配）。
    修改传入模块的所有 Font 对象。
    """
    attrs = [
        "HEADER_FONT", "BODY_FONT", "SUMMARY_FONT",
        "STATUS_FONTS", "SOURCE_FONTS",
    ]
    for attr in attrs:
        val = getattr(theme, attr, None)
        if val is None:
            continue
        if isinstance(val, dict):
            for font in val.values():
                font.name = font_name
        else:
            val.name = font_name
