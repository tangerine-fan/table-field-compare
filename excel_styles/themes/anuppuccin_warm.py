"""
AnuPpuccin Warm — 华丽铜金
深咖底 + 铜金渐变表头，厚重华丽，差异行如熔金醒目。
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAME = "anuppuccin_warm"
DESCRIPTION = "深咖底 + 铜金渐变 — 华丽厚重，差异行极醒目"

FONT_NAME = "微软雅黑"

HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color="F5E6C8")
BODY_FONT = Font(name=FONT_NAME, size=11)
SUMMARY_FONT = Font(name=FONT_NAME, bold=True, size=11, color="C0A880")

HEADER_FILL = PatternFill(start_color="8B6914", end_color="6B4E0A", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="2A1E16", end_color="2A1E16", fill_type="solid")

STATUS_FILLS = {
    "完全一致": PatternFill(start_color="2A4A2A", end_color="2A4A2A", fill_type="solid"),
    "存在差异": PatternFill(start_color="4A1A1A", end_color="4A1A1A", fill_type="solid"),
    "DEV缺失": PatternFill(start_color="4A3A10", end_color="4A3A10", fill_type="solid"),
    "标准化缺失": PatternFill(start_color="4A3A10", end_color="4A3A10", fill_type="solid"),
}

STATUS_FONTS = {
    "完全一致": Font(name=FONT_NAME, color="7EC87E"),
    "存在差异": Font(name=FONT_NAME, color="F08080"),
    "DEV缺失": Font(name=FONT_NAME, color="F0C040"),
    "标准化缺失": Font(name=FONT_NAME, color="F0C040"),
}

SOURCE_FILLS = {
    "两边共有": PatternFill(start_color="2A4A2A", end_color="2A4A2A", fill_type="solid"),
    "仅标准化有": PatternFill(start_color="4A1A1A", end_color="4A1A1A", fill_type="solid"),
    "仅DEV有": PatternFill(start_color="4A3A10", end_color="4A3A10", fill_type="solid"),
}

SOURCE_FONTS = {
    "两边共有": Font(name=FONT_NAME, color="7EC87E"),
    "仅标准化有": Font(name=FONT_NAME, color="F08080"),
    "仅DEV有": Font(name=FONT_NAME, color="F0C040"),
}

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SUMMARY_COL_WIDTHS = [25, 35, 15, 15, 15, 80]
DETAIL_COL_WIDTHS = [25, 35, 30, 15]
