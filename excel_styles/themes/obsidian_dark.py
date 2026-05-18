"""
Obsidian Dark — 暗色主题
深黑底 + 紫色表头，高对比状态色，差异行极醒目。
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAME = "obsidian_dark"
DESCRIPTION = "深黑底 + 紫色表头 — 高对比，差异行最醒目"

FONT_NAME = "微软雅黑"

HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=11)
SUMMARY_FONT = Font(name=FONT_NAME, bold=True, size=11)

HEADER_FILL = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="585B70", end_color="585B70", fill_type="solid")

STATUS_FILLS = {
    "完全一致": PatternFill(start_color="1A3A2A", end_color="1A3A2A", fill_type="solid"),
    "存在差异": PatternFill(start_color="3A1A2A", end_color="3A1A2A", fill_type="solid"),
    "DEV缺失": PatternFill(start_color="3A3510", end_color="3A3510", fill_type="solid"),
    "标准化缺失": PatternFill(start_color="3A3510", end_color="3A3510", fill_type="solid"),
}

STATUS_FONTS = {
    "完全一致": Font(name=FONT_NAME, color="A6E3A1"),
    "存在差异": Font(name=FONT_NAME, color="F38BA8"),
    "DEV缺失": Font(name=FONT_NAME, color="F9E2AF"),
    "标准化缺失": Font(name=FONT_NAME, color="F9E2AF"),
}

SOURCE_FILLS = {
    "两边共有": PatternFill(start_color="1A3A2A", end_color="1A3A2A", fill_type="solid"),
    "仅标准化有": PatternFill(start_color="3A1A2A", end_color="3A1A2A", fill_type="solid"),
    "仅DEV有": PatternFill(start_color="3A3510", end_color="3A3510", fill_type="solid"),
}

SOURCE_FONTS = {
    "两边共有": Font(name=FONT_NAME, color="A6E3A1"),
    "仅标准化有": Font(name=FONT_NAME, color="F38BA8"),
    "仅DEV有": Font(name=FONT_NAME, color="F9E2AF"),
}

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SUMMARY_COL_WIDTHS = [25, 35, 15, 15, 15, 80]
DETAIL_COL_WIDTHS = [25, 35, 30, 15]
