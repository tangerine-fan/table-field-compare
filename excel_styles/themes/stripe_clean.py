"""
Stripe Clean — 极简白底
白底 + 靛蓝强调线，SaaS 级干净感，打印友好。
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAME = "stripe_clean"
DESCRIPTION = "白底 + 靛蓝强调线 — SaaS 级干净感，打印友好"

FONT_NAME = "微软雅黑"

HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1A1F36")
BODY_FONT = Font(name=FONT_NAME, size=11)
SUMMARY_FONT = Font(name=FONT_NAME, bold=True, size=11, color="3C4257")

HEADER_FILL = PatternFill(start_color="F7F8FA", end_color="F7F8FA", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="FAFBFC", end_color="FAFBFC", fill_type="solid")

STATUS_FILLS = {
    "完全一致": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "存在差异": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "DEV缺失": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    "标准化缺失": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
}

STATUS_FONTS = {
    "完全一致": Font(name=FONT_NAME, color="1B7A2B"),
    "存在差异": Font(name=FONT_NAME, color="C62828"),
    "DEV缺失": Font(name=FONT_NAME, color="B76D00"),
    "标准化缺失": Font(name=FONT_NAME, color="B76D00"),
}

SOURCE_FILLS = {
    "两边共有": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "仅标准化有": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "仅DEV有": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
}

SOURCE_FONTS = {
    "两边共有": Font(name=FONT_NAME, color="1B7A2B"),
    "仅标准化有": Font(name=FONT_NAME, color="C62828"),
    "仅DEV有": Font(name=FONT_NAME, color="B76D00"),
}

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="E3E8EE"),
    right=Side(style="thin", color="E3E8EE"),
    top=Side(style="thin", color="E3E8EE"),
    bottom=Side(style="thin", color="E3E8EE"),
)

SUMMARY_COL_WIDTHS = [25, 35, 15, 15, 15, 80]
DETAIL_COL_WIDTHS = [25, 35, 30, 15]
