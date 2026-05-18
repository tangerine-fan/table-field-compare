"""
Catppuccin Latte — 柔和浅色
暖白底 + 马卡龙状态色，柔和护眼，打印极友好。
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAME = "catppuccin_latte"
DESCRIPTION = "暖白底 + 马卡龙色 — 柔和护眼，打印极友好"

FONT_NAME = "微软雅黑"

HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color="4C4F69")
BODY_FONT = Font(name=FONT_NAME, size=11)
SUMMARY_FONT = Font(name=FONT_NAME, bold=True, size=11, color="5C5F77")

HEADER_FILL = PatternFill(start_color="CCD0DA", end_color="CCD0DA", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E6E9EF", end_color="E6E9EF", fill_type="solid")

STATUS_FILLS = {
    "完全一致": PatternFill(start_color="A6E3A1", end_color="A6E3A1", fill_type="solid"),
    "存在差异": PatternFill(start_color="F38BA8", end_color="F38BA8", fill_type="solid"),
    "DEV缺失": PatternFill(start_color="F9E2AF", end_color="F9E2AF", fill_type="solid"),
    "标准化缺失": PatternFill(start_color="F9E2AF", end_color="F9E2AF", fill_type="solid"),
}

STATUS_FONTS = {
    "完全一致": Font(name=FONT_NAME, color="40A02B"),
    "存在差异": Font(name=FONT_NAME, color="D20F39"),
    "DEV缺失": Font(name=FONT_NAME, color="DF8E1D"),
    "标准化缺失": Font(name=FONT_NAME, color="DF8E1D"),
}

SOURCE_FILLS = {
    "两边共有": PatternFill(start_color="A6E3A1", end_color="A6E3A1", fill_type="solid"),
    "仅标准化有": PatternFill(start_color="F38BA8", end_color="F38BA8", fill_type="solid"),
    "仅DEV有": PatternFill(start_color="F9E2AF", end_color="F9E2AF", fill_type="solid"),
}

SOURCE_FONTS = {
    "两边共有": Font(name=FONT_NAME, color="40A02B"),
    "仅标准化有": Font(name=FONT_NAME, color="D20F39"),
    "仅DEV有": Font(name=FONT_NAME, color="DF8E1D"),
}

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SUMMARY_COL_WIDTHS = [25, 35, 15, 15, 15, 80]
DETAIL_COL_WIDTHS = [25, 35, 30, 15]
