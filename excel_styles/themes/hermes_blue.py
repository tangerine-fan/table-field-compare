"""
Hermes Blue — 默认主题
企业蓝表头 + 高饱和状态色，稳重现代。
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAME = "hermes_blue"
DESCRIPTION = "企业蓝表头 + 绿/红/黄状态色 — 稳重现代（默认）"

FONT_NAME = "微软雅黑"

# ── 字体 ──
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=11)
SUMMARY_FONT = Font(name=FONT_NAME, bold=True, size=11)

# ── 填充 ──
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

STATUS_FILLS = {
    "完全一致": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "存在差异": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "DEV缺失": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "标准化缺失": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

STATUS_FONTS = {
    "完全一致": Font(name=FONT_NAME, color="006100"),
    "存在差异": Font(name=FONT_NAME, color="9C0006"),
    "DEV缺失": Font(name=FONT_NAME, color="9C6500"),
    "标准化缺失": Font(name=FONT_NAME, color="9C6500"),
}

SOURCE_FILLS = {
    "两边共有": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "仅标准化有": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "仅DEV有": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}

SOURCE_FONTS = {
    "两边共有": Font(name=FONT_NAME, color="006100"),
    "仅标准化有": Font(name=FONT_NAME, color="9C0006"),
    "仅DEV有": Font(name=FONT_NAME, color="9C6500"),
}

# ── 对齐 ──
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(vertical="center", wrap_text=True)

# ── 边框 ──
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── 列宽（汇总 Sheet） ──
SUMMARY_COL_WIDTHS = [25, 35, 15, 15, 15, 80]

# ── 列宽（字段详情 Sheet） ──
DETAIL_COL_WIDTHS = [25, 35, 30, 15]
